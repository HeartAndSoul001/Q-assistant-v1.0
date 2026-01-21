from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex
from typing import List, Any
from openpyxl import load_workbook

class variableModel(QAbstractTableModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._data = []
        self._headers = []

    def rowCount(self, parent=QModelIndex()) -> int:
        return len(self._data)

    def columnCount(self, parent=QModelIndex()) -> int:
        return len(self._headers) if self._headers else 0

    def data(self, index: QModelIndex, role=Qt.DisplayRole):
        if not index.isValid():
            return None

        if role == Qt.DisplayRole:
            if index.row() < len(self._data) and index.column() < len(self._data[index.row()]):
                return str(self._data[index.row()][index.column()])
            return None
        elif role == Qt.TextAlignmentRole:
            # 设置居中对齐
            return Qt.AlignCenter
        else:
            return None

    def headerData(self, section: int, orientation: Qt.Orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal and self._headers:
                return str(self._headers[section])
            return str(section + 1)
        return None

    def setData(self, data: List[List[Any]], headers: List[str]):
        """设置新的数据和表头"""
        self.beginResetModel()
        self._data = data
        self._headers = headers
        self.endResetModel()

    def load_data_from_xlsx(self, filepath):
        try:
            workbook = load_workbook(filepath)
            sheet = workbook.active
        except Exception as e:
            print(f"Error loading data: {e}")
            return

        # 读取表头
        headers = [cell.value for cell in sheet[1]]

        # 读取数据
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(list(row))

        # 更新模型数据
        self.setData(data, headers)

    def appendRow(self, row_data: List[Any]):
        """添加新的一行数据"""
        if not self._headers or len(row_data) != len(self._headers):
            raise ValueError("数据列数与表头数不匹配")
        
        row_count = self.rowCount()
        self.beginInsertRows(QModelIndex(), row_count, row_count)
        self._data.append(row_data)
        self.endInsertRows()
    
    def get_data(self, row, column):
        # 获取指定行和列的数据
        index = self.createIndex(row, column)
        return self.data(index, Qt.DisplayRole)

    def clearData(self):
        """清空所有数据"""
        self.beginResetModel()
        self._data = []
        self._headers = []
        self.endResetModel()
